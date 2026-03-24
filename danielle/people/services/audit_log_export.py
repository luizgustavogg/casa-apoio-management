import json
from io import BytesIO


def _format_json(value):
    if not value:
        return "—"
    return json.dumps(value, ensure_ascii=False, sort_keys=True, indent=2)


def build_audit_log_xlsx(logs, filter_summary=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl nao esta instalado. Instale a dependencia para exportar XLSX."
        ) from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Historico"

    blue = PatternFill(fill_type="solid", fgColor="2F6CC2")
    light_blue = PatternFill(fill_type="solid", fgColor="EAF1FB")
    gray = PatternFill(fill_type="solid", fgColor="F8FAFC")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="2F6CC2", bold=True, size=18)
    section_font = Font(color="334155", bold=True, size=11)
    border = Border(
        left=Side(style="thin", color="D9E2EF"),
        right=Side(style="thin", color="D9E2EF"),
        top=Side(style="thin", color="D9E2EF"),
        bottom=Side(style="thin", color="D9E2EF"),
    )

    worksheet.column_dimensions["A"].width = 22
    worksheet.column_dimensions["B"].width = 14
    worksheet.column_dimensions["C"].width = 20
    worksheet.column_dimensions["D"].width = 12
    worksheet.column_dimensions["E"].width = 30
    worksheet.column_dimensions["F"].width = 45
    worksheet.column_dimensions["G"].width = 45

    worksheet.merge_cells("A1:G1")
    worksheet["A1"] = "HISTORICO DE AUDITORIA"
    worksheet["A1"].font = title_font

    worksheet.merge_cells("A2:G2")
    worksheet["A2"] = filter_summary or "Casa de Apoio"
    worksheet["A2"].font = section_font

    headers = [
        "Data",
        "Acao",
        "Entidade",
        "Registro",
        "Campos alterados",
        "Antes",
        "Depois",
    ]

    header_row = 4
    for index, text in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=index, value=text.upper())
        cell.fill = blue
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    row = header_row + 1
    for log in logs:
        values = [
            log.created_at.strftime("%d/%m/%Y %H:%M"),
            log.get_action_display(),
            log.entity,
            f"#{log.object_id}",
            ", ".join(log.changed_fields or []) or "—",
            _format_json(log.before_data),
            _format_json(log.after_data),
        ]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row, column=column, value=value)
            cell.border = border
            cell.fill = light_blue if row % 2 == 0 else gray
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
