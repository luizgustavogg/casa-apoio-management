from datetime import date, timedelta
from io import BytesIO
from statistics import mean

from django.db.models import Q

from people.models import Checkin, Checkout, HomeServices, HouseConfiguration, Person


REPORT_TYPE_OPTIONS = [
    ("occupancy", "Ocupacao e Capacidade"),
    ("people", "Pessoas"),
    ("checkin", "Check-ins"),
    ("checkout", "Check-outs"),
    ("home_services", "Utilizacoes"),
    ("gender_mix", "Mistura de Generos"),
]


def get_report_type_options():
    return REPORT_TYPE_OPTIONS


def _normalize_report_type(report_type):
    valid_types = {value for value, _ in REPORT_TYPE_OPTIONS}
    return report_type if report_type in valid_types else "occupancy"


def _build_all_days(start_date, end_date):
    total_days = (end_date - start_date).days + 1
    return [start_date + timedelta(days=offset) for offset in range(total_days)]


def _occupancy_on_day(day):
    return Checkin.objects.filter(created_at__date__lte=day).filter(
        Q(checkout__isnull=True) | Q(checkout__created_at__date__gt=day)
    ).count()


def _build_occupancy_typed_report(start_date, end_date):
    config = HouseConfiguration.get_config()
    max_capacity = config.max_capacity
    all_days = _build_all_days(start_date, end_date)

    daily = []
    for day in all_days:
        occupancy = _occupancy_on_day(day)
        rate = (occupancy / max_capacity * 100) if max_capacity > 0 else 0
        daily.append(
            {
                "date": day.isoformat(),
                "primary": occupancy,
                "secondary": round(rate, 2),
            }
        )

    primary_values = [item["primary"] for item in daily]
    avg_primary = round(mean(primary_values), 2) if primary_values else 0
    peak_primary = max(primary_values) if primary_values else 0
    min_primary = min(primary_values) if primary_values else 0
    avg_rate = round((avg_primary / max_capacity * 100), 2) if max_capacity > 0 else 0
    days_with_value = len([value for value in primary_values if value > 0])
    peak_rate = round((peak_primary / max_capacity * 100), 2) if max_capacity > 0 else 0

    peak_days = [item["date"] for item in daily if item["primary"] == peak_primary][:5]
    lowest_days = [item["date"] for item in daily if item["primary"] == min_primary][:5]

    return {
        "report_type": "occupancy",
        "title": "Relatorio de Ocupacao e Capacidade",
        "period": {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "days": len(all_days),
        },
        "summary": [
            {"label": "Capacidade maxima", "value": max_capacity, "unit": "vagas"},
            {"label": "Taxa media", "value": avg_rate, "unit": "%"},
            {"label": "Maior ocupacao", "value": peak_primary, "unit": "pessoas"},
            {"label": "Menor ocupacao", "value": min_primary, "unit": "pessoas"},
            {"label": "Dias com ocupacao", "value": days_with_value, "unit": "dias"},
            {"label": "Uso no pico", "value": peak_rate, "unit": "%"},
            {"label": "Datas de pico", "value": ", ".join(peak_days) if peak_days else "--", "unit": ""},
            {
                "label": "Datas de menor ocupacao",
                "value": ", ".join(lowest_days) if lowest_days else "--",
                "unit": "",
            },
        ],
        "daily": daily,
        "table": {
            "headers": ["Data", "Ocupacao", "Taxa (%)"],
            "keys": ["date", "primary", "secondary"],
        },
        "chart": {
            "y_axis_label": "Pessoas",
            "datasets": [
                {
                    "label": "Ocupacao diaria",
                    "data": primary_values,
                    "border_color": "#2f80ed",
                    "background_color": "rgba(47, 128, 237, 0.20)",
                    "fill": True,
                }
            ],
            "max_value_hint": max(max_capacity, max(primary_values) if primary_values else 0),
        },
    }


def _build_people_typed_report(start_date, end_date):
    all_days = _build_all_days(start_date, end_date)
    running_total = 0
    daily = []

    for day in all_days:
        count = Person.objects.filter(created_at__date=day).count()
        running_total += count
        daily.append({"date": day.isoformat(), "primary": count, "secondary": running_total})

    total_people = Person.objects.count()
    primary_values = [item["primary"] for item in daily]
    peak = max(primary_values) if primary_values else 0
    avg = round(mean(primary_values), 2) if primary_values else 0

    return {
        "report_type": "people",
        "title": "Relatorio de Pessoas",
        "period": {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "days": len(all_days),
        },
        "summary": [
            {"label": "Pessoas cadastradas no periodo", "value": sum(primary_values), "unit": "pessoas"},
            {"label": "Media diaria de cadastros", "value": avg, "unit": "pessoas"},
            {"label": "Maior dia de cadastro", "value": peak, "unit": "pessoas"},
            {"label": "Total geral de pessoas", "value": total_people, "unit": "pessoas"},
        ],
        "daily": daily,
        "table": {
            "headers": ["Data", "Cadastros", "Acumulado no periodo"],
            "keys": ["date", "primary", "secondary"],
        },
        "chart": {
            "y_axis_label": "Cadastros",
            "datasets": [
                {
                    "label": "Pessoas por dia",
                    "data": primary_values,
                    "border_color": "#2f80ed",
                    "background_color": "rgba(47, 128, 237, 0.20)",
                    "fill": True,
                }
            ],
            "max_value_hint": max(primary_values) if primary_values else 0,
        },
    }


def _build_checkin_typed_report(start_date, end_date):
    all_days = _build_all_days(start_date, end_date)
    daily = []

    for day in all_days:
        created_today = Checkin.objects.filter(created_at__date=day).count()
        active_on_day = _occupancy_on_day(day)
        daily.append({"date": day.isoformat(), "primary": created_today, "secondary": active_on_day})

    primary_values = [item["primary"] for item in daily]
    active_values = [item["secondary"] for item in daily]

    return {
        "report_type": "checkin",
        "title": "Relatorio de Check-ins",
        "period": {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "days": len(all_days),
        },
        "summary": [
            {"label": "Check-ins no periodo", "value": sum(primary_values), "unit": "eventos"},
            {"label": "Media diaria de check-ins", "value": round(mean(primary_values), 2) if primary_values else 0, "unit": "eventos"},
            {"label": "Pico diario de check-ins", "value": max(primary_values) if primary_values else 0, "unit": "eventos"},
            {"label": "Media de ativos no periodo", "value": round(mean(active_values), 2) if active_values else 0, "unit": "pessoas"},
        ],
        "daily": daily,
        "table": {
            "headers": ["Data", "Check-ins do dia", "Ativos no dia"],
            "keys": ["date", "primary", "secondary"],
        },
        "chart": {
            "y_axis_label": "Check-ins",
            "datasets": [
                {
                    "label": "Check-ins por dia",
                    "data": primary_values,
                    "border_color": "#2f80ed",
                    "background_color": "rgba(47, 128, 237, 0.20)",
                    "fill": True,
                },
                {
                    "label": "Ativos no dia",
                    "data": active_values,
                    "border_color": "#16a34a",
                    "background_color": "rgba(22, 163, 74, 0.15)",
                    "fill": False,
                },
            ],
            "max_value_hint": max(primary_values + active_values) if (primary_values or active_values) else 0,
        },
    }


def _build_checkout_typed_report(start_date, end_date):
    all_days = _build_all_days(start_date, end_date)
    daily = []
    running = 0

    for day in all_days:
        count = Checkout.objects.filter(created_at__date=day).count()
        running += count
        daily.append({"date": day.isoformat(), "primary": count, "secondary": running})

    primary_values = [item["primary"] for item in daily]

    return {
        "report_type": "checkout",
        "title": "Relatorio de Check-outs",
        "period": {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "days": len(all_days),
        },
        "summary": [
            {"label": "Check-outs no periodo", "value": sum(primary_values), "unit": "eventos"},
            {"label": "Media diaria de check-outs", "value": round(mean(primary_values), 2) if primary_values else 0, "unit": "eventos"},
            {"label": "Pico diario de check-outs", "value": max(primary_values) if primary_values else 0, "unit": "eventos"},
        ],
        "daily": daily,
        "table": {
            "headers": ["Data", "Check-outs do dia", "Acumulado no periodo"],
            "keys": ["date", "primary", "secondary"],
        },
        "chart": {
            "y_axis_label": "Check-outs",
            "datasets": [
                {
                    "label": "Check-outs por dia",
                    "data": primary_values,
                    "border_color": "#ea580c",
                    "background_color": "rgba(234, 88, 12, 0.18)",
                    "fill": True,
                }
            ],
            "max_value_hint": max(primary_values) if primary_values else 0,
        },
    }


def _home_services_score_for_day(day):
    services = HomeServices.objects.filter(created_at__date=day)
    rows = services.values(
        "breakfast", "lunch", "snack", "dinner", "shower", "sleep"
    )
    total_actions = 0
    for item in rows:
        total_actions += sum(1 for value in item.values() if value)
    return services.count(), total_actions


def _build_home_services_typed_report(start_date, end_date):
    all_days = _build_all_days(start_date, end_date)
    daily = []
    for day in all_days:
        records, actions = _home_services_score_for_day(day)
        daily.append({"date": day.isoformat(), "primary": records, "secondary": actions})

    primary_values = [item["primary"] for item in daily]
    action_values = [item["secondary"] for item in daily]

    return {
        "report_type": "home_services",
        "title": "Relatorio de Utilizacoes",
        "period": {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "days": len(all_days),
        },
        "summary": [
            {"label": "Registros no periodo", "value": sum(primary_values), "unit": "registros"},
            {"label": "Acoes de servico no periodo", "value": sum(action_values), "unit": "acoes"},
            {"label": "Media diaria de registros", "value": round(mean(primary_values), 2) if primary_values else 0, "unit": "registros"},
            {"label": "Pico diario de acoes", "value": max(action_values) if action_values else 0, "unit": "acoes"},
        ],
        "daily": daily,
        "table": {
            "headers": ["Data", "Registros", "Acoes de servico"],
            "keys": ["date", "primary", "secondary"],
        },
        "chart": {
            "y_axis_label": "Utilizacoes",
            "datasets": [
                {
                    "label": "Registros por dia",
                    "data": primary_values,
                    "border_color": "#0891b2",
                    "background_color": "rgba(8, 145, 178, 0.20)",
                    "fill": True,
                },
                {
                    "label": "Acoes por dia",
                    "data": action_values,
                    "border_color": "#0f766e",
                    "background_color": "rgba(15, 118, 110, 0.15)",
                    "fill": False,
                },
            ],
            "max_value_hint": max(primary_values + action_values) if (primary_values or action_values) else 0,
        },
    }


def _build_gender_mix_typed_report(start_date, end_date):
    all_days = _build_all_days(start_date, end_date)
    daily = []

    for day in all_days:
        male = Person.objects.filter(created_at__date=day, gender="M").count()
        female = Person.objects.filter(created_at__date=day, gender="F").count()
        other = Person.objects.filter(created_at__date=day, gender="O").count()
        total = male + female + other
        daily.append(
            {
                "date": day.isoformat(),
                "male": male,
                "female": female,
                "other": other,
                "total": total,
            }
        )

    male_total = Person.objects.filter(gender="M").count()
    female_total = Person.objects.filter(gender="F").count()
    other_total = Person.objects.filter(gender="O").count()
    unspecified_total = Person.objects.filter(gender__isnull=True).count() + Person.objects.filter(gender="").count()

    return {
        "report_type": "gender_mix",
        "title": "Relatorio de Mistura de Generos",
        "period": {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "days": len(all_days),
        },
        "summary": [
            {"label": "Masculino", "value": male_total, "unit": "pessoas"},
            {"label": "Feminino", "value": female_total, "unit": "pessoas"},
            {"label": "Outro", "value": other_total, "unit": "pessoas"},
            {"label": "Nao informado", "value": unspecified_total, "unit": "pessoas"},
            {"label": "Total geral", "value": male_total + female_total + other_total + unspecified_total, "unit": "pessoas"},
        ],
        "daily": daily,
        "table": {
            "headers": ["Data", "Masculino", "Feminino", "Outro", "Total"],
            "keys": ["date", "male", "female", "other", "total"],
        },
        "chart": {
            "y_axis_label": "Cadastros",
            "datasets": [
                {
                    "label": "Masculino",
                    "data": [item["male"] for item in daily],
                    "border_color": "#2563eb",
                    "background_color": "rgba(37, 99, 235, 0.18)",
                    "fill": False,
                },
                {
                    "label": "Feminino",
                    "data": [item["female"] for item in daily],
                    "border_color": "#ec4899",
                    "background_color": "rgba(236, 72, 153, 0.18)",
                    "fill": False,
                },
                {
                    "label": "Outro",
                    "data": [item["other"] for item in daily],
                    "border_color": "#7c3aed",
                    "background_color": "rgba(124, 58, 237, 0.18)",
                    "fill": False,
                },
            ],
            "max_value_hint": max([item["total"] for item in daily]) if daily else 0,
        },
    }


def build_typed_report(start_date, end_date, report_type="occupancy"):
    if start_date > end_date:
        start_date, end_date = end_date, start_date

    normalized = _normalize_report_type(report_type)

    if normalized == "people":
        return _build_people_typed_report(start_date, end_date)
    if normalized == "checkin":
        return _build_checkin_typed_report(start_date, end_date)
    if normalized == "checkout":
        return _build_checkout_typed_report(start_date, end_date)
    if normalized == "home_services":
        return _build_home_services_typed_report(start_date, end_date)
    if normalized == "gender_mix":
        return _build_gender_mix_typed_report(start_date, end_date)
    return _build_occupancy_typed_report(start_date, end_date)


def build_occupancy_report(start_date, end_date):
    """Build occupancy/capacity metrics for a date interval (inclusive)."""
    if start_date > end_date:
        start_date, end_date = end_date, start_date

    config = HouseConfiguration.get_config()
    max_capacity = config.max_capacity

    total_days = (end_date - start_date).days + 1
    all_days = [start_date + timedelta(days=offset) for offset in range(total_days)]

    daily = []
    for day in all_days:
        occupancy = Checkin.objects.filter(
            created_at__date__lte=day
        ).filter(
            Q(checkout__isnull=True) | Q(checkout__created_at__date__gt=day)
        ).count()

        occupancy_rate = (occupancy / max_capacity * 100) if max_capacity > 0 else 0
        daily.append(
            {
                "date": day.isoformat(),
                "occupancy": occupancy,
                "occupancy_rate": round(occupancy_rate, 2),
            }
        )

    occupancies = [item["occupancy"] for item in daily]
    avg_occupancy = mean(occupancies) if occupancies else 0
    avg_occupancy_rate = (avg_occupancy / max_capacity * 100) if max_capacity > 0 else 0

    max_occupancy = max(occupancies) if occupancies else 0
    min_occupancy = min(occupancies) if occupancies else 0

    peak_days = [item for item in daily if item["occupancy"] == max_occupancy]
    lowest_days = [item for item in daily if item["occupancy"] == min_occupancy]

    days_with_occupancy = len([value for value in occupancies if value > 0])
    capacity_used_peak_rate = (max_occupancy / max_capacity * 100) if max_capacity > 0 else 0

    return {
        "period": {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "days": total_days,
        },
        "capacity": {
            "max_capacity": max_capacity,
            "avg_occupancy": round(avg_occupancy, 2),
            "avg_occupancy_rate": round(avg_occupancy_rate, 2),
            "peak_occupancy": max_occupancy,
            "lowest_occupancy": min_occupancy,
            "days_with_occupancy": days_with_occupancy,
            "capacity_used_peak_rate": round(capacity_used_peak_rate, 2),
        },
        "peak_days": peak_days,
        "lowest_days": lowest_days,
        "daily": daily,
    }



def build_occupancy_report_pdf(html_content):
    """Convert report HTML content to PDF bytes."""
    try:
        from xhtml2pdf import pisa
    except ImportError as exc:
        raise RuntimeError(
            "xhtml2pdf nao esta instalado. Instale a dependencia para exportar PDF."
        ) from exc

    output = BytesIO()
    pdf_status = pisa.CreatePDF(src=html_content, dest=output, encoding="utf-8")

    if pdf_status.err:
        raise RuntimeError("Falha ao gerar PDF a partir do template HTML.")

    return output.getvalue()


def build_occupancy_report_xlsx(report_data):
    """Build an XLSX report with styled sections and occupancy chart."""
    if "title" not in report_data:
        # Backward compatibility for callers still using build_occupancy_report().
        start_date = date.fromisoformat(report_data["period"]["start_date"])
        end_date = date.fromisoformat(report_data["period"]["end_date"])
        report_data = build_typed_report(start_date, end_date, report_type="occupancy")

    try:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl nao esta instalado. Instale a dependencia para exportar XLSX."
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"

    blue = PatternFill(fill_type="solid", fgColor="2F6CC2")
    light_blue = PatternFill(fill_type="solid", fgColor="EAF1FB")
    gray = PatternFill(fill_type="solid", fgColor="F5F7FA")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="2F6CC2", bold=True, size=18)
    subtitle_font = Font(color="334155", bold=True, size=11)
    section_font = Font(color="2F6CC2", bold=True, size=12)
    border = Border(
        left=Side(style="thin", color="D9E2EF"),
        right=Side(style="thin", color="D9E2EF"),
        top=Side(style="thin", color="D9E2EF"),
        bottom=Side(style="thin", color="D9E2EF"),
    )

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    ws.merge_cells("A1:E1")
    ws["A1"] = report_data["title"].upper()
    ws["A1"].font = title_font

    ws.merge_cells("A2:D2")
    ws["A2"] = "Casa de Apoio"
    ws["A2"].font = subtitle_font
    ws["E2"] = report_data["period"]["end_date"][:4]
    ws["E2"].alignment = Alignment(horizontal="right")
    ws["E2"].font = Font(color="64748B", size=11)

    ws.merge_cells("A4:E4")
    ws["A4"] = "METRICAS-CHAVE"
    ws["A4"].font = section_font

    metric_cards = report_data["summary"][:5]

    for col, item in enumerate(metric_cards, start=1):
        label = item["label"].upper()
        value = item["value"]
        unit = item.get("unit", "")
        header_cell = ws.cell(row=6, column=col, value=label)
        header_cell.fill = blue
        header_cell.font = white_font
        header_cell.alignment = Alignment(horizontal="center")
        header_cell.border = border

        value_cell = ws.cell(row=7, column=col, value=value)
        value_cell.font = Font(color="334155", size=16, bold=True)
        value_cell.fill = gray
        value_cell.alignment = Alignment(horizontal="center")
        value_cell.border = border

        unit_cell = ws.cell(row=8, column=col, value=unit)
        unit_cell.font = Font(color="64748B", size=10)
        unit_cell.fill = gray
        unit_cell.alignment = Alignment(horizontal="center")
        unit_cell.border = border

    ws.merge_cells("A10:E10")
    ws["A10"] = "TODAS AS METRICAS"
    ws["A10"].font = section_font

    header_row = 12
    headers = ["METRICA", "VALOR", "UNIDADE", "REFERENCIA", "OBS"]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.fill = blue
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    summary_rows = [
        ("Dias no periodo", report_data["period"]["days"], "dias", "", ""),
    ]
    for item in report_data["summary"]:
        summary_rows.append((item["label"], item["value"], item.get("unit", ""), "", ""))

    row = header_row + 1
    for values in summary_rows:
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.fill = light_blue if row % 2 == 0 else gray
            if col in (2, 3):
                cell.alignment = Alignment(horizontal="center")
        row += 1

    row += 1
    daily_header_row = row
    table_headers = report_data.get("table", {}).get("headers", ["Data", "Valor", "Valor 2"])
    daily_headers = [header.upper() for header in table_headers[:5]]
    while len(daily_headers) < 5:
        daily_headers.append("")
    for col, text in enumerate(daily_headers, start=1):
        cell = ws.cell(row=daily_header_row, column=col, value=text)
        cell.fill = blue
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    daily_start = daily_header_row + 1
    current = daily_start
    table_keys = report_data.get("table", {}).get("keys", ["date", "primary", "secondary"])
    for item in report_data["daily"]:
        values = [item.get(key, "") for key in table_keys[:5]]
        while len(values) < 5:
            values.append("")
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=current, column=col, value=value)
            cell.border = border
            cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF" if current % 2 else "F8FAFC")
            if col in (2, 3):
                cell.alignment = Alignment(horizontal="center")
        current += 1

    if current > daily_start:
        chart = LineChart()
        chart.title = "Tendencia"
        chart.y_axis.title = report_data.get("chart", {}).get("y_axis_label", "Valor")
        chart.x_axis.title = "Data"
        chart.height = 8
        chart.width = 14
        chart.style = 2

        categories = Reference(ws, min_col=1, min_row=daily_start, max_row=current - 1)
        max_series_col = min(len(table_keys), 4)
        data = Reference(ws, min_col=2, max_col=max_series_col + 1, min_row=daily_start - 1, max_row=current - 1)
        chart.add_data(data, titles_from_data=True, from_rows=False)
        chart.set_categories(categories)
        if max_series_col == 1:
            chart.legend = None
        ws.add_chart(chart, f"A{current + 2}")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
